# -*-coding:utf-8 -*-
import io
import os
import openpyxl
def case_avg(input_path,i,thread_num):
    print('clean audio')
    audio_list = os.listdir('./result/audio/')
    for fileName in audio_list:
        os.remove('./result/audio/' + fileName)
    print('clean done')

    file_list = os.listdir('./avg_tmp/')
    for fileName in file_list:
        os.remove('./avg_tmp/' + fileName)

    case = []
    '''
    case_data = xlrd.open_workbook(input_path)
    table = case_data.sheets()[i]
    list_case = table.col_values(0, start_rowx=1, end_rowx=None)  # 语料
    for z in range(len(list_case)):
        cell = table.cell(z+1, 0)
        if cell.ctype == 2 and cell.value % 1 == 0:
            cell_value = int(cell.value)
            #print(cell_value)
            case.append(cell_value)
    '''
    case_data = openpyxl.load_workbook(input_path)
    #sheetnames = case_data.get_sheet_names()
    table = case_data.worksheets[i]
    rows = table.max_row
    for z in range(1,rows):
        cell = table.cell(z + 1, 1).value
        if cell:
            #print(cell)
            case.append(cell)

    list_case = [x for x in case if x != '']
    line_num_list = len(list_case)
    if line_num_list < int(thread_num):
        print("并发数大于用例总数,程序退出")
        exit()


    # 根据并发数分配每个进程的用例数
    case_avg_path = './avg_tmp/'
    if not os.path.exists(case_avg_path):
        os.mkdir(case_avg_path)
    avg_tmp_list = os.listdir("./avg_tmp/")
    for i in range(len(avg_tmp_list)):
        os.remove(case_avg_path + avg_tmp_list[i])
    if line_num_list % int(thread_num) == 0:
        avg_num = line_num_list / int(thread_num)
        remainder = 0
    else:
        avg_num = line_num_list // int(thread_num)
        remainder = line_num_list % int(thread_num)

    case_list = []

    for i in range(int(thread_num)):
        if remainder == 0:
            case_list.append(int(avg_num))
        else:
            if i < remainder:
                case_list.append(int(avg_num) + 1)
            else:
                case_list.append(int(avg_num))
    # 将每个进程的用例数,写入文件
    min = 0
    for i in range(len(case_list)):
        if i == 0:
            avg_list = list_case[0:case_list[i]]
        else:
            min += case_list[i - 1]
            max = min + case_list[i]
            avg_list = list_case[min:max]

        #log_tmp.error(avg_list)
        avg_file = io.open(case_avg_path + str(i), 'w+', encoding='UTF-8')
        for xx in range(len(avg_list)):
            if xx == len(avg_list) - 1:
                avg_file.write(str(avg_list[xx]))
            else:
                avg_file.write(str(avg_list[xx]) + '\n')
            #log_tmp.error(str(avg_list[xx]))
        avg_file.close()
    return len(list_case)
