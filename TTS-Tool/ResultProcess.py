#-*- coding=utf-8 -*-

import time
import openpyxl
class ResultProcess:
    def __init__(self,i):
        self.run_time = time.strftime('%Y-%m-%d.%H.%M.%S', time.localtime(time.time()))
        self.open_excel = openpyxl.Workbook()
        self.open_sheet = self.open_excel.create_sheet('sheet' + str(i), i)
        self.open_sheet['A1'] = '测试语料'
        self.open_sheet['B1'] = '音频结果'

    def SaveResult(self,ret_list):
        #print("保存结果中...")
        case_text_ls = []
        audiores_ls = []

        run_time = self.run_time
        open_excel = self.open_excel
        open_sheet = self.open_sheet

        for i in range(len(ret_list)):
            result = ret_list[i]
            case_text_ls.append(result[0])
            audiores_ls.append(result[1])

        if len(case_text_ls) > 0:
            for z in range(len(case_text_ls)):
                open_sheet.cell(z + 2, 1, case_text_ls[z])
                open_sheet.cell(z + 2, 2, audiores_ls[z])
            open_excel.save('./result/' + 'TTS_result_' + str(run_time) +  '.xlsx')
        #open_sheet.write(x + 1, 4, result)
        #print("结果保存完成")