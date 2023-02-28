# -*- coding: utf-8 -*-
# @Time    : 2019/12/16 13:50
# @Author  : huangyl
# @FileName: ListenSong_test.py
# @Software: PyCharm
# @Blog    ：https://hardelm.com

from tools.log4j import *
from readConfig import ReadConfig
import requests
import json
import jsonpath
import xlwt
import xlrd
import openpyxl
import itertools
import time
import os
import io
from multiprocessing import Pool, Manager
import linecache

localReadConfig = ReadConfig()
url = localReadConfig.get_url('url')
appkey = localReadConfig.get_header('appkey')
udid = localReadConfig.get_header('udid')
ver = localReadConfig.get_header('ver')
method = localReadConfig.get_header('method')
appsig = localReadConfig.get_header('appsig')
#case_path = localReadConfig.get_case_path('case_path')
result_path = localReadConfig.get_result_path('result_path')

def case_avg(input_path,i,thread_num):
    file_list = os.listdir('./avg_tmp/')
    for fileName in file_list:
        os.remove('./avg_tmp/' + fileName)

    case = []
    '''
    case_data = xlrd.open_workbook(input_path)
    table = case_data.sheets()[i]
    list_case = table.col_values(0, start_rowx=1, end_rowx=None)  # 语料
    for z in range(len(list_case)):
        cell = table.cell(z+1, 0)
        if cell.ctype == 2 and cell.value % 1 == 0:
            cell_value = int(cell.value)
            #print(cell_value)
            case.append(cell_value)
    '''
    case_data = openpyxl.load_workbook(input_path)
    #sheetnames = case_data.get_sheet_names()
    table = case_data.worksheets[i]
    rows = table.max_row
    for z in range(1,rows):
        cell = table.cell(z + 1, 1).value
        if cell:
            #print(cell)
            case.append(cell)

    list_case = [x for x in case if x != '']
    line_num_list = len(list_case)
    if line_num_list < int(thread_num):
        print("并发数大于用例总数,程序退出")
        exit()


    # 根据并发数分配每个进程的用例数
    case_avg_path = './avg_tmp/'
    if not os.path.exists(case_avg_path):
        os.mkdir(case_avg_path)
    avg_tmp_list = os.listdir("./avg_tmp/")
    for i in range(len(avg_tmp_list)):
        os.remove(case_avg_path + avg_tmp_list[i])
    if line_num_list % int(thread_num) == 0:
        avg_num = line_num_list / int(thread_num)
        remainder = 0
    else:
        avg_num = line_num_list // int(thread_num)
        remainder = line_num_list % int(thread_num)

    case_list = []

    for i in range(int(thread_num)):
        if remainder == 0:
            case_list.append(int(avg_num))
        else:
            if i < remainder:
                case_list.append(int(avg_num) + 1)
            else:
                case_list.append(int(avg_num))
    # 将每个进程的用例数,写入文件
    min = 0
    for i in range(len(case_list)):
        if i == 0:
            avg_list = list_case[0:case_list[i]]
        else:
            min += case_list[i - 1]
            max = min + case_list[i]
            avg_list = list_case[min:max]

        #log_tmp.error(avg_list)
        avg_file = io.open(case_avg_path + str(i), 'w+', encoding='UTF-8')
        for xx in range(len(avg_list)):
            if xx == len(avg_list) - 1:
                avg_file.write(str(avg_list[xx]))
            else:
                avg_file.write(str(avg_list[xx]) + '\n')
            #log_tmp.error(str(avg_list[xx]))
        avg_file.close()
    return len(list_case)


def generateText(case,i):
    if i == 0:
        text = '我要听' + case + '的歌'
        return text
    elif i == 1:
        text = '我要听' + case
        return text

def generateNLUresult(file_name,case_res,i):
    list_tmp = linecache.getlines('./avg_tmp/' + file_name)
    for z in range(len(list_tmp)):
        #print(case[z])
        testcase = list_tmp[z].strip('\n')
        if isinstance(testcase,float):
            testcase = int(testcase)

        #print(testcase)
        if i == 0:
            text = '我要听' + str(testcase) + '的歌'
            full_url = (url + '?appkey=' + appkey + '&udid=' + udid + '&ver=' + ver + '&method=' + method + '&appsig=' + appsig + '&text=' + text)
        elif i == 1:
            text = '我要听' + str(testcase)
            full_url = (url + '?appkey=' + appkey + '&udid=' + udid + '&ver=' + ver + '&method=' + method + '&appsig=' + appsig + '&text=' + text)
        print(full_url)
        try:
            ret = requests.get(full_url).text
            #print(str(ret))
        except Exception as e:
            print(e)
            ret = '请求异常！'
        case_res.append([list_tmp[z],ret])
        #case_res.append([caseText, ret])
        #print(case_res)
    return case_res

def actProcess(case_res):
    resultlist = []
    for i in range(len(case_res)):
        case = case_res[i]
        textS = case[0]
        ret = case[1]
        if isinstance(textS,float):
            textS = str(int(textS))
        testLower = textS.lower()
        try:
            act_res = json.loads(ret)
            service = act_res["service"]
            if service == "cn.yunzhisheng.music":
                if act_res["code"] == "SEARCH_SONG" or act_res["code"] == "SEARCH_ARTIST":
                    if jsonpath.jsonpath(act_res, '$..keyword'):
                        keyword = act_res["semantic"]["intent"]["keyword"]
                        #print("keyword:" + keyword)
                        #print("textS:" + textS)
                        #print("testLower:" + testLower)
                        if keyword in testLower:
                            if jsonpath.jsonpath(act_res, '$..artist'):
                                artist = act_res["semantic"]["intent"]["artist"]
                                #print(artist)
                                if artist in testLower:
                                    resultlist.append([textS, ret, "pass"])
                                else:
                                    resultlist.append([textS, ret, "checkartist"])
                            elif jsonpath.jsonpath(act_res, '$..song'):
                                song = act_res["semantic"]["intent"]["song"]
                                #print(song)
                                if song in testLower:
                                    resultlist.append([textS, ret, "pass"])
                                else:
                                    resultlist.append([textS, ret, "check song"])
                            else:
                                resultlist.append([textS, ret, "artist|song not exsit"])
                        else:
                            resultlist.append([textS, ret, "check keyword"])
                    else:
                        resultlist.append([textS, ret, "keyword not exsit"])
                else:
                    resultlist.append([textS, ret, "code不匹配"])
            else:
                resultlist.append([textS, ret, "service不匹配"])

        except BaseException as err:
            print(err)
    return resultlist

def resultProcess(resultlist,i):
    #print(resultlist)
    case_text = []
    res = []
    Ans = []
    for z in range(len(resultlist)):
        result = resultlist[z]
        case_text.append(result[0])
        res.append(result[1])
        Ans.append(result[2])

    run_time = time.strftime('%Y-%m-%d.%H.%M.%S', time.localtime(time.time()))
    '''
    open_excel = xlwt.Workbook(encoding='utf-8', style_compression=0)
    open_sheet = open_excel.add_sheet('Sheet'+str(i))
    open_sheet.write(0, 0, u'语料')
    open_sheet.write(0, 1, u'实际res')
    open_sheet.write(0, 2, u'结果')
    if len(case_text) > 0:
        for x in range(len(case_text)):
            open_sheet.write(x+1, 0, case_text[x])
            open_sheet.write(x+1, 1, res[x])
            open_sheet.write(x+1, 2, Ans[x])
    '''
    open_excel = openpyxl.Workbook()
    open_sheet = open_excel.create_sheet('sheet'+str(i),i)
    open_sheet['A1'] = '语料'
    open_sheet['B1'] = '实际res'
    open_sheet['C1'] = '结果'

    if len(case_text) > 0:
        for x in range(len(case_text)):
            open_sheet.cell(x+2, 1, case_text[x])
            open_sheet.cell(x+2, 2, res[x])
            open_sheet.cell(x+2, 3, Ans[x])
        open_excel.save('./result/' + 'NLU_result_' + str(run_time) +  '.xlsx')
'''
def main(x):
    for i in range(x):
        list_case = getcase(case_path, i)
        case_res = generateNLUresult(list_case, i)
        resultlist = actProcess(case_res)
        resultProcess(resultlist,i)
    #return resultlist
'''

def main(list_res,i):
    file_list = os.listdir("./avg_tmp/")
    try:
        p = Pool(len(file_list))
        # log_tmp.error(len(file_list))
        for z in range(len(file_list)):
            p.apply_async(generateNLUresult, args=(file_list[z],list_res,i))
        p.close()
        p.join()
    except Exception as err_msg:
        print("main():error mesage=%s" % str(err_msg))
    return list_res


if __name__ == '__main__':
    case_path = './case/case.xlsx'
    i = 0
    #resultlist = main()
    #resultProcess(resultlist)
    all_num = case_avg(case_path,i,2)
    list_tmp = Manager().list()
    main(list_tmp,i)
    resultlist = actProcess(list_tmp)
    resultProcess(resultlist,i)









